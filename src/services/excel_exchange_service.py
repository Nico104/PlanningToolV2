from __future__ import annotations

import json
import re
from pathlib import Path
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .termin_occurrence_service import SUPPORTED_PERIODIZITAET, series_date_sequence
from .semester_rules import semester_from_id


_FILE_SCHEMAS: Dict[str, Dict[str, Any]] = {
    "raeume.json": {
        "sheet": "Raeume",
        "list_key": "raeume",
        "columns": ["id", "name", "kapazitaet", "gebaeude"],
    },
    "lehrveranstaltungen.json": {
        "sheet": "Lehrveranstaltungen",
        "list_key": "lehrveranstaltungen",
        "columns": [
            "id",
            "name",
            "vortragende.name",
            "vortragende.email",
            "studiensemester",
            "studienrichtung",
            "ects",
        ],
    },
    "termine.json": {
        "sheet": "Termine",
        "list_key": "termine",
        "columns": [
            "id",
            "name",
            "lva_id",
            "typ",
            "datum",
            "datum_bis",
            "periodizitaet",
            "start_zeit",
            "raum_id",
            "gruppe.name",
            "gruppe.groesse",
            "anwesenheitspflicht",
            "duration",
            "semester_id",
            "ausfall_daten",
            "serien_ausnahmen",
            "notiz",
            "zu_besprechen",
            "besprechungshinweis",
        ],
    },
    "studienrichtungen.json": {
        "sheet": "Studienrichtungen",
        "list_key": "studienrichtungen",
        "columns": ["id", "name"],
    },
    "freie_tage.json": {
        "sheet": "FreieTage",
        "list_key": "freie_tage",
        "columns": ["id", "typ", "beschreibung", "datum", "von_datum", "bis_datum"],
    },
}

_EXCEL_HEADER_LABELS: Dict[str, str] = {
    "raeume.json:id": "Raumnummer",
    "raeume.json:name": "Raum",
    "raeume.json:kapazitaet": "Kapazität",
    "raeume.json:gebaeude": "Gebäude",
    "lehrveranstaltungen.json:id": "LVA-Nr.",
    "lehrveranstaltungen.json:name": "Name",
    "lehrveranstaltungen.json:vortragende.name": "Vortragende",
    "lehrveranstaltungen.json:vortragende.email": "E-Mail",
    "lehrveranstaltungen.json:studienrichtung": "Studienrichtung",
    "lehrveranstaltungen.json:studiensemester": "Studiensemester",
    "lehrveranstaltungen.json:ects": "ECTS",
    "termine.json:lva_id": "LVA-Nr.",
    "termine.json:zu_besprechen": "Zu besprechen",
    "termine.json:besprechungshinweis": "Hinweis",
}

_EXCEL_HEADER_ALIASES: Dict[str, Dict[str, str]] = {
    "lehrveranstaltungen.json": {
        "LVA-Nr": "id",
        "LVA-Nummer": "id",
        "LVA": "id",
        "LVA-Name": "name",
        "Lehrveranstaltung": "name",
        "Vortragende": "vortragende.name",
        "Vortragende Name": "vortragende.name",
        "Lehrperson": "vortragende.name",
        "E-Mail": "vortragende.email",
        "Email": "vortragende.email",
        "Studiensemester": "studiensemester",
        "Studienrichtung": "studienrichtung",
        "ECTS": "ects",
    },
    "raeume.json": {
        "Raumnummer": "id",
        "Raum-Nr": "id",
        "Raum": "name",
        "Kapazität": "kapazitaet",
        "Kapazitaet": "kapazitaet",
        "Gebäude": "gebaeude",
        "Gebaeude": "gebaeude",
        "Gebäudekürzel": "gebaeude",
        "Gebaeudekuerzel": "gebaeude",
    },
    "termine.json": {
        "LVA-Nr": "lva_id",
        "LVA-Nummer": "lva_id",
        "Zu besprechen": "zu_besprechen",
        "ZuBesprechen": "zu_besprechen",
        "Besprechen": "zu_besprechen",
        "Hinweis": "besprechungshinweis",
        "Besprechungshinweis": "besprechungshinweis",
    },
}


@dataclass(frozen=True)
class TeacherExportOption:
    name: str
    email: str
    lva_count: int
    term_count: int
    semester_term_counts: Dict[str, int] = field(default_factory=dict)
    semester_lva_ids: Dict[str, tuple[str, ...]] = field(default_factory=dict)

    @property
    def key(self) -> tuple[str, str]:
        return (self.name, self.email)

    def counts_for_semesters(self, semester_ids: Optional[Iterable[str]]) -> tuple[int, int]:
        if semester_ids is None:
            return self.lva_count, self.term_count

        selected_ids = {_safe_text(item) for item in semester_ids}
        lva_ids: set[str] = set()
        term_count = 0
        for semester_id in selected_ids:
            term_count += int(self.semester_term_counts.get(semester_id, 0))
            lva_ids.update(self.semester_lva_ids.get(semester_id, ()))
        return len(lva_ids), term_count


@dataclass(frozen=True)
class SemesterExportOption:
    id: str
    name: str
    term_count: int


@dataclass(frozen=True)
class LvaExportOption:
    id: str
    name: str
    teacher_name: str
    teacher_email: str
    term_count: int
    semester_term_counts: Dict[str, int] = field(default_factory=dict)

    def counts_for_semesters(self, semester_ids: Optional[Iterable[str]]) -> tuple[int, int]:
        if semester_ids is None:
            return 1, self.term_count

        selected_ids = {_safe_text(item) for item in semester_ids}
        term_count = sum(int(self.semester_term_counts.get(semester_id, 0)) for semester_id in selected_ids)
        return (1 if term_count > 0 else 0), term_count


def _read_json_file(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8-sig"))


def _read_json_list(path: Path, list_key: str) -> List[Dict[str, Any]]:
    payload = _read_json_file(path)
    rows = payload.get(list_key, []) if isinstance(payload, dict) else []
    return rows if isinstance(rows, list) else []


def _semester_export_rows(data_dir: Path) -> List[Dict[str, Any]]:
    extra_ids = {
        _safe_text(item.get("semester_id"))
        for item in _read_json_list(data_dir / "termine.json", "termine")
        if _safe_text(item.get("semester_id"))
    }
    rows = []
    for semester_id in extra_ids:
        semester = semester_from_id(semester_id)
        if semester is None:
            rows.append({"id": semester_id, "name": semester_id, "start": "", "end": ""})
            continue
        rows.append(
            {
                "id": semester.id,
                "name": semester.name,
                "start": semester.start.isoformat(),
                "end": semester.end.isoformat(),
            }
        )
    return sorted(rows, key=lambda item: (item.get("start") or "9999-12-31", item.get("id") or ""))


def _semester_display_name(semester_id: str) -> str:
    semester_id = _safe_text(semester_id)
    if not semester_id:
        return "Ohne Semester"
    semester = semester_from_id(semester_id)
    return semester.name if semester else semester_id


def _semester_sort_key(semester_id: str) -> tuple:
    semester_id = _safe_text(semester_id)
    if not semester_id:
        return (date.max, "zzzz")
    semester = semester_from_id(semester_id)
    if semester:
        return (semester.start, semester.id)
    return (date.max, semester_id)


def _get_nested(data: Dict[str, Any], path: str) -> Any:
    node: Any = data
    for part in path.split("."):
        if not isinstance(node, dict):
            return None
        node = node.get(part)
        if node is None:
            return None
    return node


def _set_nested(data: Dict[str, Any], path: str, value: Any) -> None:
    parts = path.split(".")
    node = data
    for part in parts[:-1]:
        nxt = node.get(part)
        if not isinstance(nxt, dict):
            nxt = {}
            node[part] = nxt
        node = nxt
    node[parts[-1]] = value


def _serialize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, list):
        if any(isinstance(item, dict) for item in value):
            return json.dumps(value, ensure_ascii=False)
        return ";".join(str(v) for v in value)
    if isinstance(value, bool):
        return "Ja" if value else "Nein"
    return value


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _normalize_excel_header(value: Any) -> str:
    text = _safe_text(value).casefold()
    for src, repl in {
        "ä": "ae",
        "ö": "oe",
        "ü": "ue",
        "ß": "ss",
    }.items():
        text = text.replace(src, repl)
    return re.sub(r"[^a-z0-9]+", "", text)


def _column_from_excel_header(file_name: str, header: Any) -> Optional[str]:
    normalized = _normalize_excel_header(header)
    if not normalized:
        return None

    cfg = _FILE_SCHEMAS.get(file_name, {})
    columns = list(cfg.get("columns", []))
    lookup: Dict[str, str] = {}
    for col in columns:
        lookup[_normalize_excel_header(col)] = col
        label = _EXCEL_HEADER_LABELS.get(f"{file_name}:{col}")
        if label:
            lookup[_normalize_excel_header(label)] = col

    for label, col in _EXCEL_HEADER_ALIASES.get(file_name, {}).items():
        if col in columns:
            lookup[_normalize_excel_header(label)] = col

    return lookup.get(normalized)


def _safe_date(value: Any) -> Optional[date]:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = _safe_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            continue
    return None


def _safe_time(value: Any) -> Optional[time]:
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    text = _safe_text(value)
    if not text:
        return None
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).time()
        except Exception:
            continue
    return None


def _safe_end_time(start_value: Any, duration_value: Any) -> Optional[time]:
    start = _safe_time(start_value)
    duration = _parse_int(duration_value)
    if start is None or duration is None or duration <= 0:
        return None
    base = datetime.combine(date(2000, 1, 1), start)
    return (base + timedelta(minutes=duration)).time()


def _excel_compatible_sheet_name(base_name: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[\\/?*\[\]:]", "", _safe_text(base_name))
    cleaned = re.sub(r"\s+", " ", cleaned).strip() or "Ohne Lehrperson"
    cleaned = cleaned[:31]

    candidate = cleaned
    index = 2
    while candidate in used_names:
        suffix = f" {index}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}".strip()
        index += 1
    used_names.add(candidate)
    return candidate


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    txt = str(value).strip().lower()
    return txt in {"1", "true", "wahr", "yes", "ja"}


def _parse_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    txt = str(value).strip()
    if not txt:
        return None
    return int(float(txt))


def _parse_list(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    txt = str(value).strip()
    if not txt:
        return []
    return [p.strip() for p in txt.split(";") if p.strip()]


def _parse_series_exceptions(value: Any) -> List[Dict[str, Any]]:
    if value is None or value == "":
        return []
    raw = value
    if isinstance(value, str):
        try:
            raw = json.loads(value)
        except Exception:
            return []
    if not isinstance(raw, list):
        return []

    out: List[Dict[str, Any]] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        original = _safe_date(item.get("original_datum"))
        target = _safe_date(item.get("datum"))
        if original is None or target is None:
            continue
        start = _safe_time(item.get("start_zeit"))
        duration = _parse_int(item.get("duration"))
        normalized: Dict[str, Any] = {
            "original_datum": original.isoformat(),
            "datum": target.isoformat(),
            "start_zeit": start.strftime("%H:%M") if start else None,
            "raum_id": _safe_text(item.get("raum_id")) or None,
            "duration": duration,
        }
        out.append(normalized)
    return out


def _series_dates_from_entry(termin: Dict[str, Any]) -> List[date]:
    start = _safe_date(termin.get("datum"))
    end = _safe_date(termin.get("datum_bis"))
    period = _safe_text(termin.get("periodizitaet"))
    if start is None:
        return []
    if end is None or end < start or period not in SUPPORTED_PERIODIZITAET:
        return [start]
    return series_date_sequence(start, end, period)


def _series_exceptions_by_original_date(termin: Dict[str, Any]) -> Dict[date, Dict[str, Any]]:
    out: Dict[date, Dict[str, Any]] = {}
    for item in _parse_series_exceptions(termin.get("serien_ausnahmen")):
        original = _safe_date(item.get("original_datum"))
        if original is not None:
            out[original] = item
    return out


def _expand_termin_entries(termine: Iterable[Dict[str, Any]]) -> Iterable[Dict[str, Any]]:
    for termin in termine:
        if not isinstance(termin, dict):
            continue
        dates = _series_dates_from_entry(termin)
        if not dates:
            yield termin
            continue
        skipped = {d for d in (_safe_date(item) for item in _parse_list(termin.get("ausfall_daten"))) if d is not None}
        exceptions = _series_exceptions_by_original_date(termin)
        if len(dates) == 1:
            item = dict(termin)
            original = dates[0]
            if original in skipped:
                continue
            exception = exceptions.get(original)
            item["datum"] = (exception.get("datum") if exception else original.isoformat())
            if exception:
                if exception.get("start_zeit"):
                    item["start_zeit"] = exception.get("start_zeit")
                if exception.get("raum_id"):
                    item["raum_id"] = exception.get("raum_id")
                if exception.get("duration") is not None:
                    item["duration"] = exception.get("duration")
            yield item
            continue
        for original_date in dates:
            if original_date in skipped:
                continue
            item = dict(termin)
            exception = exceptions.get(original_date)
            item["datum"] = exception.get("datum") if exception else original_date.isoformat()
            if exception:
                if exception.get("start_zeit"):
                    item["start_zeit"] = exception.get("start_zeit")
                if exception.get("raum_id"):
                    item["raum_id"] = exception.get("raum_id")
                if exception.get("duration") is not None:
                    item["duration"] = exception.get("duration")
            yield item


def _normalize_entry(file_name: str, entry: Dict[str, Any]) -> Dict[str, Any]:
    if file_name == "raeume.json":
        if "kapazitaet" in entry:
            val = _parse_int(entry.get("kapazitaet"))
            entry["kapazitaet"] = 0 if val is None else val
    elif file_name == "lehrveranstaltungen.json":
        entry.pop("typ", None)
        entry["studiensemester"] = _parse_list(entry.get("studiensemester"))
    elif file_name == "termine.json":
        entry["notiz"] = str(entry.get("notiz", ""))
        entry["zu_besprechen"] = _parse_bool(entry.get("zu_besprechen"))
        entry["besprechungshinweis"] = str(entry.get("besprechungshinweis", ""))
        entry["raum_id"] = str(entry.get("raum_id", ""))
        entry["semester_id"] = str(entry.get("semester_id", ""))
        period = str(entry.get("periodizitaet", "") or "").strip()
        entry["periodizitaet"] = None if not period or period.lower() == "keine" else period
        entry["ausfall_daten"] = _parse_list(entry.get("ausfall_daten"))
        entry["serien_ausnahmen"] = _parse_series_exceptions(entry.get("serien_ausnahmen"))

        if "duration" in entry:
            val = _parse_int(entry.get("duration"))
            entry["duration"] = 0 if val is None else val
        else:
            entry["duration"] = 0

        entry["anwesenheitspflicht"] = _parse_bool(entry.get("anwesenheitspflicht"))

        gruppe = entry.get("gruppe")
        if isinstance(gruppe, dict):
            g_name = str(gruppe.get("name", "")).strip()
            g_size = _parse_int(gruppe.get("groesse"))
            if g_name or g_size not in (None, 0):
                entry["gruppe"] = {"name": g_name or "-", "groesse": g_size or 0}
            else:
                entry["gruppe"] = None
        else:
            entry["gruppe"] = None

        if entry.get("datum") in (None, ""):
            entry["datum"] = None
        if entry.get("datum_bis") in (None, ""):
            entry["datum_bis"] = None
            entry["periodizitaet"] = None
        if entry.get("start_zeit") in (None, ""):
            entry["start_zeit"] = None
    elif file_name == "freie_tage.json":
        if not str(entry.get("id", "")).strip():
            entry.pop("id", None)
        datum = str(entry.get("datum", "")).strip()
        von = str(entry.get("von_datum", "")).strip()
        bis = str(entry.get("bis_datum", "")).strip()
        if datum:
            entry["datum"] = datum
            entry.pop("von_datum", None)
            entry.pop("bis_datum", None)
        elif von and bis:
            entry["von_datum"] = von
            entry["bis_datum"] = bis
            entry.pop("datum", None)
    return entry


def export_project_to_excel(data_dir: Path, output_path: Path) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for file_name, cfg in _FILE_SCHEMAS.items():
        sheet = wb.create_sheet(cfg["sheet"])
        columns: List[str] = cfg["columns"]
        list_key = cfg["list_key"]
        sheet.append([_EXCEL_HEADER_LABELS.get(f"{file_name}:{col}", col) for col in columns])

        payload = _read_json_file(data_dir / file_name)
        rows = payload.get(list_key, []) if isinstance(payload, dict) else []
        if not isinstance(rows, list):
            rows = []

        for item in rows:
            if not isinstance(item, dict):
                continue
            row = [_serialize_cell(_get_nested(item, col)) for col in columns]
            sheet.append(row)

    wb.save(output_path)


def export_week_calendar_to_excel(
    data_dir: Path,
    output_path: Path,
    date_from: date,
    date_to: Optional[date] = None,
    teacher_filter: Optional[Iterable[tuple[str, str]]] = None,
    semester_filter: Optional[Iterable[str]] = None,
    lva_filter: Optional[Iterable[str]] = None,
    include_weekend: bool = False,
    slot_minutes: int = 30,
) -> None:
    if date_to is None:
        date_to = date_from + timedelta(days=6 - date_from.weekday())
    if date_to < date_from:
        date_from, date_to = date_to, date_from

    teacher_filter_set = (
        {(_safe_text(item[0]), _safe_text(item[1] if len(item) > 1 else "")) for item in teacher_filter}
        if teacher_filter is not None
        else None
    )
    semester_filter_set = (
        {_safe_text(item) for item in semester_filter}
        if semester_filter is not None
        else None
    )
    lva_filter_set = (
        {_safe_text(item) for item in lva_filter if _safe_text(item)}
        if lva_filter is not None
        else None
    )
    lvas = _read_json_list(data_dir / "lehrveranstaltungen.json", "lehrveranstaltungen")
    termine = _read_json_list(data_dir / "termine.json", "termine")
    raeume = _read_json_list(data_dir / "raeume.json", "raeume")

    lva_map = {_safe_text(item.get("id")): item for item in lvas if _safe_text(item.get("id"))}
    raum_map = {_safe_text(item.get("id")): item for item in raeume if _safe_text(item.get("id"))}
    lva_teacher: Dict[str, tuple[str, str]] = {}
    for lva in lvas:
        if not isinstance(lva, dict):
            continue
        lva_id = _safe_text(lva.get("id"))
        if lva_id:
            lva_teacher[lva_id] = (
                _safe_text(lva.get("vortragende", {}).get("name")),
                _safe_text(lva.get("vortragende", {}).get("email")),
            )

    rows_by_week: Dict[date, List[Dict[str, Any]]] = defaultdict(list)
    for termin in _expand_termin_entries(termine):
        if not isinstance(termin, dict):
            continue
        if semester_filter_set is not None and _safe_text(termin.get("semester_id")) not in semester_filter_set:
            continue
        lva_id = _safe_text(termin.get("lva_id"))
        if lva_filter_set is not None and lva_id not in lva_filter_set:
            continue

        datum = _safe_date(termin.get("datum"))
        start = _safe_time(termin.get("start_zeit"))
        duration = _parse_int(termin.get("duration")) or 45
        if datum is None or start is None or datum < date_from or datum > date_to:
            continue
        if not include_weekend and datum.weekday() >= 5:
            continue

        teacher_name, teacher_email = lva_teacher.get(lva_id, ("", ""))
        if teacher_filter_set is not None and (teacher_name, teacher_email) not in teacher_filter_set:
            continue

        lva = lva_map.get(lva_id, {})
        room_id = _safe_text(termin.get("raum_id"))
        room = raum_map.get(room_id, {})
        gruppe = termin.get("gruppe") if isinstance(termin.get("gruppe"), dict) else {}
        start_minutes = start.hour * 60 + start.minute
        week_start = datum - timedelta(days=datum.weekday())
        rows_by_week[week_start].append(
            {
                "date": datum,
                "start": start,
                "start_minutes": start_minutes,
                "end_minutes": max(start_minutes + duration, start_minutes + 15),
                "duration": duration,
                "lva_id": lva_id,
                "lva_name": _safe_text(lva.get("name")) if isinstance(lva, dict) else _safe_text(termin.get("name")),
                "room": _safe_text(room.get("name")) if isinstance(room, dict) else room_id,
                "gruppe": _safe_text(gruppe.get("name")) if isinstance(gruppe, dict) else "",
                "typ": _safe_text(termin.get("typ")),
                "teacher": teacher_name,
                "zu_besprechen": _parse_bool(termin.get("zu_besprechen")),
                "besprechungshinweis": _safe_text(termin.get("besprechungshinweis")),
            }
        )

    wb = Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()
    sheet = wb.create_sheet(_excel_compatible_sheet_name("Wochenkalender", used_names))

    if not rows_by_week:
        week_start = date_from - timedelta(days=date_from.weekday())
        _write_week_calendar_sheet(
            sheet,
            [],
            week_start,
            "Wochenkalender",
            include_weekend=include_weekend,
            slot_minutes=slot_minutes,
        )
    else:
        start_row = 1
        for week_start, rows in sorted(rows_by_week.items(), key=lambda item: item[0]):
            start_row = _write_week_calendar_sheet(
                sheet,
                rows,
                week_start,
                "Wochenkalender",
                include_weekend=include_weekend,
                slot_minutes=slot_minutes,
                start_row=start_row,
            )

    wb.save(output_path)


def _write_week_calendar_sheet(
    sheet,
    rows: List[Dict[str, Any]],
    week_start: date,
    owner_name: str,
    *,
    include_weekend: bool,
    slot_minutes: int,
    start_row: int = 1,
) -> int:
    slot_minutes = 60 if int(slot_minutes) == 60 else 30
    day_count = 7 if include_weekend else 5
    week_end = week_start + timedelta(days=day_count - 1)
    weekday_names = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"][:day_count]
    type_colors = {
        "VO": "E3F2FD",
        "UE": "E8F5E9",
        "LU": "FFF3E0",
        "SE": "F3E5F5",
    }
    fallback_colors = ["F7F7F7", "DDEBF7", "E2F0D9", "FFF2CC", "E4DFEC"]

    min_minutes = 8 * 60
    max_minutes = 20 * 60
    rows_by_day: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        datum = _safe_date(row.get("date"))
        if datum is None:
            continue
        if datum.weekday() >= day_count:
            continue
        min_minutes = min(min_minutes, (int(row["start_minutes"]) // slot_minutes) * slot_minutes)
        max_minutes = max(
            max_minutes,
            ((int(row["end_minutes"]) + slot_minutes - 1) // slot_minutes) * slot_minutes,
        )
        rows_by_day[datum.weekday()].append(row)

    day_lane_counts: Dict[int, int] = {}
    for day_index, day_rows in rows_by_day.items():
        lane_end_minutes: List[int] = []
        for item in sorted(day_rows, key=lambda row: (int(row["start_minutes"]), int(row["end_minutes"]), row["lva_id"])):
            start_minutes = int(item["start_minutes"])
            end_minutes = int(item["end_minutes"])
            lane_index = next(
                (index for index, lane_end in enumerate(lane_end_minutes) if lane_end <= start_minutes),
                None,
            )
            if lane_index is None:
                lane_index = len(lane_end_minutes)
                lane_end_minutes.append(end_minutes)
            else:
                lane_end_minutes[lane_index] = end_minutes
            item["_lane"] = lane_index
        day_lane_counts[day_index] = max(1, len(lane_end_minutes))

    day_column_starts: Dict[int, int] = {}
    next_column = 2
    for day_index in range(day_count):
        day_column_starts[day_index] = next_column
        next_column += day_lane_counts.get(day_index, 1)

    min_minutes = max(0, min_minutes)
    max_minutes = min(24 * 60, max(max_minutes, min_minutes + 60))
    slot_count = ((max_minutes - min_minutes) + slot_minutes - 1) // slot_minutes
    row_height = 24 if slot_minutes == 30 else 30

    title = f"KW {week_start.isocalendar().week} ({week_start:%d.%m.%Y} - {week_end:%d.%m.%Y})"
    last_column = next_column - 1
    title_row = start_row
    header_row = start_row + 1
    first_slot_row = start_row + 2
    sheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=last_column)
    sheet.cell(row=title_row, column=1).value = title
    sheet.cell(row=title_row, column=1).font = Font(bold=True, size=12, color="1F2933")
    sheet.cell(row=title_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[title_row].height = 22

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2EC")
    time_fill = PatternFill("solid", fgColor="F7F9FB")
    block_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF"),
    )
    discuss_border = Border(
        left=Side(style="medium", color="D98200"),
        right=Side(style="medium", color="D98200"),
        top=Side(style="medium", color="D98200"),
        bottom=Side(style="medium", color="D98200"),
    )

    sheet.cell(row=header_row, column=1).value = "Zeit"
    for col in range(1, last_column + 1):
        cell = sheet.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color="9EADBA"))

    for day_index, day_name in enumerate(weekday_names):
        current = week_start + timedelta(days=day_index)
        start_column = day_column_starts[day_index]
        end_column = start_column + day_lane_counts.get(day_index, 1) - 1
        if end_column > start_column:
            sheet.merge_cells(start_row=header_row, start_column=start_column, end_row=header_row, end_column=end_column)
        sheet.cell(row=header_row, column=start_column).value = f"{day_name}\n{current:%d.%m.%Y}"

    for slot_index in range(slot_count + 1):
        minutes = min_minutes + slot_index * slot_minutes
        row = first_slot_row + slot_index
        sheet.cell(row=row, column=1).value = f"{minutes // 60:02d}:{minutes % 60:02d}"
        sheet.cell(row=row, column=1).fill = time_fill
        sheet.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
        sheet.cell(row=row, column=1).font = Font(size=9, color="52616B")
        sheet.row_dimensions[row].height = row_height
        for col in range(1, last_column + 1):
            sheet.cell(row=row, column=col).border = Border(
                left=thin_gray,
                right=thin_gray,
                top=thin_gray,
                bottom=thin_gray,
            )

    if not rows:
        note = sheet.cell(row=first_slot_row, column=2)
        note.value = "Keine Termine im gewählten Zeitraum."
        note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        note.font = Font(size=10, italic=True, color="52616B")

    fallback_key_colors: Dict[str, str] = {}
    for day_index, day_rows in rows_by_day.items():
        day_start_column = day_column_starts[day_index]
        for item in sorted(day_rows, key=lambda row: (row["start_minutes"], row.get("_lane", 0), row["lva_id"])):
            col = day_start_column + int(item.get("_lane", 0))
            typ_key = _safe_text(item.get("typ")).upper()
            if typ_key in type_colors:
                fill_color = type_colors[typ_key]
            else:
                fallback_key = typ_key or item["lva_id"] or item["lva_name"]
                if fallback_key not in fallback_key_colors:
                    fallback_key_colors[fallback_key] = fallback_colors[len(fallback_key_colors) % len(fallback_colors)]
                fill_color = fallback_key_colors[fallback_key]
            fill = PatternFill("solid", fgColor=fill_color)

            start_slot = (item["start_minutes"] - min_minutes) // slot_minutes
            end_slot = max(start_slot + 1, (item["end_minutes"] - min_minutes + slot_minutes - 1) // slot_minutes)
            block_start_row = first_slot_row + max(0, start_slot)
            block_end_row = first_slot_row + min(slot_count, end_slot) - 1

            for row in range(block_start_row, block_end_row + 1):
                cell = sheet.cell(row=row, column=col)
                cell.fill = fill
                cell.border = discuss_border if item.get("zu_besprechen") else block_border

            text_parts = [
                f"{item['start'].strftime('%H:%M')} - {((datetime.combine(date(2000, 1, 1), item['start']) + timedelta(minutes=item['duration'])).time()).strftime('%H:%M')}",
                item.get("teacher"),
                item["room"],
                item["lva_id"],
                item["lva_name"],
                item["gruppe"],
            ]
            text = " | ".join(part for part in text_parts if part)
            hint = _safe_text(item.get("besprechungshinweis"))
            if block_end_row > block_start_row:
                sheet.merge_cells(start_row=block_start_row, start_column=col, end_row=block_end_row, end_column=col)
            top_cell = sheet.cell(row=block_start_row, column=col)
            if item.get("zu_besprechen") and hint:
                top_cell.value = CellRichText(
                    text,
                    "\n",
                    TextBlock(InlineFont(b=True, sz=9, color="8A4D00"), f"Hinweis: {hint}"),
                )
            elif item.get("zu_besprechen"):
                top_cell.value = CellRichText(
                    text,
                    "\n",
                    TextBlock(InlineFont(b=True, sz=9, color="8A4D00"), "Zu besprechen"),
                )
            else:
                top_cell.value = f"{top_cell.value}\n{text}" if top_cell.value else text
            top_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            top_cell.font = Font(size=9, color="1F2933")

    sheet.column_dimensions["A"].width = 9
    for col in range(2, last_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 24
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    return first_slot_row + slot_count + 2


def _is_empty_row(values: Tuple[Any, ...]) -> bool:
    for val in values:
        if val is not None and str(val).strip() != "":
            return False
    return True


def import_project_from_excel(excel_path: Path) -> Dict[str, Dict[str, Any]]:
    wb = load_workbook(excel_path, data_only=True)
    result: Dict[str, Dict[str, Any]] = {}

    for file_name, cfg in _FILE_SCHEMAS.items():
        sheet_name = cfg["sheet"]
        columns: List[str] = cfg["columns"]
        list_key = cfg["list_key"]

        if sheet_name not in wb.sheetnames:
            sheet_name = next((alias for alias in cfg.get("sheet_aliases", []) if alias in wb.sheetnames), sheet_name)
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        entries: List[Dict[str, Any]] = []

        header_columns: Dict[int, str] = {}
        for col_idx in range(1, (ws.max_column or 0) + 1):
            mapped = _column_from_excel_header(file_name, ws.cell(row=1, column=col_idx).value)
            if mapped and mapped not in header_columns.values():
                header_columns[col_idx] = mapped

        if header_columns:
            row_iter = ws.iter_rows(min_row=2, max_col=ws.max_column, values_only=True)
        else:
            row_iter = ws.iter_rows(min_row=2, max_col=len(columns), values_only=True)

        for values in row_iter:
            if _is_empty_row(values):
                continue

            entry: Dict[str, Any] = {}
            if header_columns:
                mapped_columns = ((idx - 1, col) for idx, col in header_columns.items())
            else:
                mapped_columns = enumerate(columns)

            for idx, col in mapped_columns:
                raw = values[idx] if idx < len(values) else None
                if raw is None:
                    continue
                txt = str(raw).strip()
                if txt == "":
                    continue
                _set_nested(entry, col, txt)

            normalized = _normalize_entry(file_name, entry)
            entries.append(normalized)

        result[file_name] = {list_key: entries}

    return result


_TISS_ROOM_COLUMN_ALIASES: Dict[str, set[str]] = {
    "id": {"raumnummer", "raumnr", "raumnummercode", "raumcode", "code", "id", "nummer"},
    "name": {"raum", "raumname", "raumbezeichnung", "bezeichnung", "name"},
    "kapazitaet": {"kapazitaet", "kapazitat", "plaetze", "platze", "plätze", "kapaz", "capacity"},
    "gebaeude": {
        "gebaeude",
        "gebauede",
        "gebäude",
        "gebaude",
        "gebaeudekuerzel",
        "gebaudekurzel",
        "gebaeudecode",
        "gebaudecode",
        "building",
        "buildingcode",
    },
    "adresse": {"adresse", "anschrift", "address"},
}


def _normalize_tiss_header(value: Any) -> str:
    text = _safe_text(value).casefold()
    for src, repl in {
        "ä": "ae",
        "ö": "oe",
        "ü": "ue",
        "ß": "ss",
    }.items():
        text = text.replace(src, repl)
    return re.sub(r"[^a-z0-9]+", "", text)


def _find_tiss_room_columns(ws) -> tuple[int, dict[str, int]] | None:
    max_scan_row = min(ws.max_row or 0, 30)
    max_scan_col = min(ws.max_column or 0, 30)

    for row_idx in range(1, max_scan_row + 1):
        mapping: dict[str, int] = {}
        for col_idx in range(1, max_scan_col + 1):
            header = _normalize_tiss_header(ws.cell(row=row_idx, column=col_idx).value)
            if not header:
                continue
            for target, aliases in _TISS_ROOM_COLUMN_ALIASES.items():
                if target not in mapping and header in aliases:
                    mapping[target] = col_idx
                    break

        if "id" in mapping and "name" in mapping and "kapazitaet" in mapping:
            return row_idx, mapping

    return None


def import_tiss_rooms_from_excel(excel_path: Path) -> Dict[str, Dict[str, Any]]:
    wb = load_workbook(excel_path, data_only=True)
    rooms_by_id: dict[str, dict[str, Any]] = {}

    for ws in wb.worksheets:
        detected = _find_tiss_room_columns(ws)
        if detected is None:
            continue

        header_row, columns = detected
        id_col = columns["id"]
        name_col = columns["name"]
        capacity_col = columns["kapazitaet"]
        building_col = columns.get("gebaeude")
        address_col = columns.get("adresse")

        for row_idx in range(header_row + 1, (ws.max_row or 0) + 1):
            room_id = _safe_text(ws.cell(row=row_idx, column=id_col).value)
            name = _safe_text(ws.cell(row=row_idx, column=name_col).value)
            if not room_id or not name:
                continue

            capacity = _parse_int(ws.cell(row=row_idx, column=capacity_col).value)
            if capacity is None:
                continue

            room = {
                "id": room_id,
                "name": name,
                "kapazitaet": capacity,
            }
            building = _safe_text(ws.cell(row=row_idx, column=building_col).value) if building_col else ""
            address = _safe_text(ws.cell(row=row_idx, column=address_col).value) if address_col else ""
            if building:
                room["gebaeude"] = building
                room["__catalog_gebaeude"] = building
            if address:
                room["__catalog_adresse"] = address
            rooms_by_id[room_id] = room

    rooms = list(rooms_by_id.values())
    if not rooms:
        raise ValueError(
            "Keine TISS-Räume erkannt. Erwartet werden mindestens die Spalten "
            "'Raumnummer', 'Raum' und 'Kapazität'. Optional wird 'Gebäude' übernommen."
        )

    return {"raeume.json": {"raeume": rooms}}


def get_teacher_export_semester_options(data_dir: Path) -> List[SemesterExportOption]:
    termine = _read_json_list(data_dir / "termine.json", "termine")
    term_counts: Dict[str, int] = defaultdict(int)
    for termin in _expand_termin_entries(termine):
        if not isinstance(termin, dict):
            continue
        term_counts[_safe_text(termin.get("semester_id"))] += 1

    return [
        SemesterExportOption(
            id=semester_id,
            name=_semester_display_name(semester_id),
            term_count=term_counts[semester_id],
        )
        for semester_id in sorted(term_counts.keys(), key=_semester_sort_key)
    ]


def get_teacher_export_options(data_dir: Path) -> List[TeacherExportOption]:
    lvas = _read_json_list(data_dir / "lehrveranstaltungen.json", "lehrveranstaltungen")
    termine = _read_json_list(data_dir / "termine.json", "termine")

    teacher_lvas: Dict[tuple[str, str], set[str]] = defaultdict(set)
    lva_teacher: Dict[str, tuple[str, str]] = {}

    for lva in lvas:
        if not isinstance(lva, dict):
            continue
        name = _safe_text(lva.get("vortragende", {}).get("name"))
        email = _safe_text(lva.get("vortragende", {}).get("email"))
        lva_id = _safe_text(lva.get("id"))
        if not name:
            continue
        key = (name, email)
        if lva_id:
            teacher_lvas[key].add(lva_id)
            lva_teacher[lva_id] = key
        else:
            teacher_lvas[key]

    teacher_terms: Dict[tuple[str, str], int] = defaultdict(int)
    teacher_semester_terms: Dict[tuple[str, str], Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    teacher_semester_lvas: Dict[tuple[str, str], Dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    for termin in _expand_termin_entries(termine):
        if not isinstance(termin, dict):
            continue
        lva_id = _safe_text(termin.get("lva_id"))
        key = lva_teacher.get(lva_id)
        if key:
            semester_id = _safe_text(termin.get("semester_id"))
            teacher_terms[key] += 1
            teacher_semester_terms[key][semester_id] += 1
            if lva_id:
                teacher_semester_lvas[key][semester_id].add(lva_id)

    return [
        TeacherExportOption(
            name=name,
            email=email,
            lva_count=len(teacher_lvas[(name, email)]),
            term_count=teacher_terms[(name, email)],
            semester_term_counts=dict(teacher_semester_terms[(name, email)]),
            semester_lva_ids={
                semester_id: tuple(sorted(lva_ids))
                for semester_id, lva_ids in teacher_semester_lvas[(name, email)].items()
            },
        )
        for name, email in sorted(teacher_lvas.keys(), key=lambda item: (item[0].lower(), item[1].lower()))
    ]


def get_lva_export_options(data_dir: Path) -> List[LvaExportOption]:
    lvas = _read_json_list(data_dir / "lehrveranstaltungen.json", "lehrveranstaltungen")
    termine = _read_json_list(data_dir / "termine.json", "termine")

    lva_terms: Dict[str, int] = defaultdict(int)
    lva_semester_terms: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for termin in _expand_termin_entries(termine):
        if not isinstance(termin, dict):
            continue
        lva_id = _safe_text(termin.get("lva_id"))
        if not lva_id:
            continue
        semester_id = _safe_text(termin.get("semester_id"))
        lva_terms[lva_id] += 1
        lva_semester_terms[lva_id][semester_id] += 1

    options: List[LvaExportOption] = []
    seen_lva_ids: set[str] = set()
    for lva in lvas:
        if not isinstance(lva, dict):
            continue
        lva_id = _safe_text(lva.get("id"))
        if not lva_id:
            continue
        seen_lva_ids.add(lva_id)
        teacher = lva.get("vortragende", {}) if isinstance(lva.get("vortragende"), dict) else {}
        options.append(
            LvaExportOption(
                id=lva_id,
                name=_safe_text(lva.get("name")),
                teacher_name=_safe_text(teacher.get("name")),
                teacher_email=_safe_text(teacher.get("email")),
                term_count=lva_terms[lva_id],
                semester_term_counts=dict(lva_semester_terms[lva_id]),
            )
        )

    for lva_id in sorted(set(lva_terms) - seen_lva_ids):
        options.append(
            LvaExportOption(
                id=lva_id,
                name="",
                teacher_name="",
                teacher_email="",
                term_count=lva_terms[lva_id],
                semester_term_counts=dict(lva_semester_terms[lva_id]),
            )
        )

    return sorted(options, key=lambda item: (item.teacher_name.lower(), item.id.lower(), item.name.lower()))


def export_terms_for_teachers_to_excel(
    data_dir: Path,
    output_path: Path,
    teacher_filter: Optional[Iterable[tuple[str, str]]] = None,
    semester_filter: Optional[Iterable[str]] = None,
    lva_filter: Optional[Iterable[str]] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> None:
    headers = [
        "Raum",
        "Datum",
        "Beginn",
        "Ende",
        "LVA-Nr.",
        "Lehrveranstaltung",
        "Lehrperson",
        "E-Mail",
        "Typ",
        "Semester",
        "Gruppe",
        "Gruppengröße",
        "Anwesenheitspflicht",
        "Zu besprechen",
        "Hinweis",
    ]

    teacher_filter_set = (
        {(_safe_text(item[0]), _safe_text(item[1] if len(item) > 1 else "")) for item in teacher_filter}
        if teacher_filter is not None
        else None
    )
    semester_filter_set = (
        {_safe_text(item) for item in semester_filter}
        if semester_filter is not None
        else None
    )
    lva_filter_set = (
        {_safe_text(item) for item in lva_filter if _safe_text(item)}
        if lva_filter is not None
        else None
    )
    if date_from is not None and date_to is not None and date_to < date_from:
        date_from, date_to = date_to, date_from
    lvas = _read_json_list(data_dir / "lehrveranstaltungen.json", "lehrveranstaltungen")
    termine = _read_json_list(data_dir / "termine.json", "termine")
    raeume = _read_json_list(data_dir / "raeume.json", "raeume")
    semester = _semester_export_rows(data_dir)

    lva_map = {_safe_text(item.get("id")): item for item in lvas if _safe_text(item.get("id"))}
    raum_map = {_safe_text(item.get("id")): item for item in raeume if _safe_text(item.get("id"))}
    semester_map = {_safe_text(item.get("id")): item for item in semester if _safe_text(item.get("id"))}

    rows_by_teacher: Dict[tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    rows_without_teacher: List[Dict[str, Any]] = []
    if teacher_filter_set is not None:
        for teacher_key in teacher_filter_set:
            if teacher_key[0]:
                rows_by_teacher[teacher_key]

    for termin in _expand_termin_entries(termine):
        if not isinstance(termin, dict):
            continue

        if semester_filter_set is not None and _safe_text(termin.get("semester_id")) not in semester_filter_set:
            continue

        lva_id = _safe_text(termin.get("lva_id"))
        if lva_filter_set is not None and lva_id not in lva_filter_set:
            continue

        lva = lva_map.get(lva_id)
        teacher_name = _safe_text(lva.get("vortragende", {}).get("name")) if isinstance(lva, dict) else ""
        teacher_email = _safe_text(lva.get("vortragende", {}).get("email")) if isinstance(lva, dict) else ""
        if teacher_filter_set is not None and (teacher_name, teacher_email) not in teacher_filter_set:
            continue
        room_item = raum_map.get(_safe_text(termin.get("raum_id")))
        semester_item = semester_map.get(_safe_text(termin.get("semester_id")))
        gruppe = termin.get("gruppe") if isinstance(termin.get("gruppe"), dict) else {}
        datum = _safe_date(termin.get("datum"))
        start = _safe_time(termin.get("start_zeit"))
        ende = _safe_end_time(termin.get("start_zeit"), termin.get("duration"))
        if date_from is not None and (datum is None or datum < date_from):
            continue
        if date_to is not None and (datum is None or datum > date_to):
            continue

        row = {
            "LVA-Nr.": _safe_text(lva.get("id")) if isinstance(lva, dict) else _safe_text(termin.get("lva_id")),
            "Lehrveranstaltung": _safe_text(lva.get("name")) if isinstance(lva, dict) else _safe_text(termin.get("name")),
            "Lehrperson": teacher_name,
            "E-Mail": teacher_email,
            "Typ": _safe_text(termin.get("typ")),
            "Semester": _safe_text(semester_item.get("name")) if isinstance(semester_item, dict) else _safe_text(termin.get("semester_id")),
            "Gruppe": _safe_text(gruppe.get("name")) if isinstance(gruppe, dict) else "",
            "Gruppengröße": gruppe.get("groesse", "") if isinstance(gruppe, dict) else "",
            "Raum": _safe_text(room_item.get("name")) if isinstance(room_item, dict) else _safe_text(termin.get("raum_id")),
            "Datum": datum.isoformat() if datum else "",
            "Beginn": start.strftime("%H:%M") if start else "",
            "Ende": ende.strftime("%H:%M") if ende else "",
            "Anwesenheitspflicht": "Ja" if bool(termin.get("anwesenheitspflicht", False)) else "Nein",
            "Zu besprechen": "Ja" if _parse_bool(termin.get("zu_besprechen")) else "Nein",
            "Hinweis": _safe_text(termin.get("besprechungshinweis")),
        }

        (rows_by_teacher[(teacher_name, teacher_email)] if teacher_name else rows_without_teacher).append(row)

    def sort_key(row: Dict[str, Any]) -> tuple:
        datum = _safe_date(row.get("Datum"))
        start = _safe_time(row.get("Beginn"))
        return (datum is None, datum or date.max, start is None, start or time.max, _safe_text(row.get("LVA-Nr.")))

    def write_sheet(sheet, data_rows: List[Dict[str, Any]]) -> None:
        sheet.append(headers)
        for row in data_rows:
            sheet.append([row.get(col, "") for col in headers])
        if data_rows:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=False)
        border = Border(bottom=Side(style="thin", color="9EADBA"))
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_cells in sheet.columns:
            width = 0
            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue
                text = value.strftime("%H:%M") if isinstance(value, time) else value.isoformat() if isinstance(value, date) and not isinstance(value, datetime) else str(value)
                width = max(width, len(text))
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width + 2, 10), 36)

    wb = Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()

    if rows_without_teacher and teacher_filter_set is None:
        sheet = wb.create_sheet(_excel_compatible_sheet_name("Ohne Lehrperson", used_names))
        write_sheet(sheet, sorted(rows_without_teacher, key=sort_key))

    teacher_counts: Dict[str, int] = defaultdict(int)
    for (teacher_name, teacher_email), data_rows in sorted(rows_by_teacher.items(), key=lambda item: (item[0][0].lower(), item[0][1].lower())):
        if not data_rows:
            continue
        base_name = teacher_name or teacher_email or "Ohne Lehrperson"
        teacher_counts[base_name] += 1
        display_name = base_name if teacher_counts[base_name] == 1 else f"{base_name} {teacher_counts[base_name]}"
        sheet = wb.create_sheet(_excel_compatible_sheet_name(display_name, used_names))
        write_sheet(sheet, sorted(data_rows, key=sort_key))

    if not wb.sheetnames:
        sheet = wb.create_sheet(_excel_compatible_sheet_name("Keine Termine", used_names))
        write_sheet(sheet, [])

    wb.save(output_path)
